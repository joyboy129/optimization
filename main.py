import pulp
from pulp import *
import pandas as pd
import numpy as np
import openpyxl
import gurobipy
from gurobipy import *
import xlsxwriter
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
#Input


def dict(a,b):
    my={}
    for i in range(len(a)):
        my[a[i]]=b[i]
    return my
def getting(i,l):
    m=[]
    for k in l:
        m.append(k[i])
    return m
def concat_list(a,b):
    m=[]
    for i in range(len(a)):
        m.append(a[i]+"_"+b[i])
    return m

def making_recette(a,b,c):
    m={}
    n=len(b[0])
    for i in range(len(a)):
        m[a[i]]=dict(c,b[i][2:])
    return m
def elim_zero(l):
    m=[]
    for i in l:
        if i !=0 and i!="0":
            m.append(i)
    return m
def making_dict(a,b):
    m={}
    for i in range(len(a)):
        m[a[i]]=b[i]
    return m
def com_possible(a,b):
    m=[]
    for i in a:
        for j in b:
            m.append(i+"_"+str(j))
    return m
def making_pourc(df,l,MP):
    m={}
    for i in range(len(l)):
        p={}
        for j in range(len(MP)):
            p[MP[j]]=df[i][j+2]
        m[l[i]]=p
    return m
Cout_de_stockage_MP=10
Cout_de_stockage_Engrais=10
Cout_blending=120
Cout_smart_blending=260
blending_min=500
stockage_min=100
Big_M=100000000
Cout_stockage_building=10
Big_n=range(1,2)
Big_n_1=range(2,3)
url1="venv/Structure_Modele_Actuelle - small data set.xlsx"
url2="venv/Structure_Modele_Actuelle - all data set.xlsx"
url=url1
OCP=["OCP Jorf","OCP SAFI","OCP Port Casablanca","OCP Port Tanger","OCP Port Nador"]
df_Options = pd.read_excel (url, sheet_name='Options')
df_Options=df_Options.to_numpy()
Cout_blender=10000000
Cout_sblender=100000
temps=[2,3,4,1]


#Lecture des excels
df_MP = pd.read_excel (url, sheet_name='MP')
df_MP=df_MP.to_numpy()
df_Stockage = pd.read_excel (url, sheet_name='Stockage')
df_Stockage=df_Stockage.to_numpy()
df_Blenders = pd.read_excel (url, sheet_name='Blenders')
df_Blenders=df_Blenders.to_numpy()
df_Recette = pd.read_excel (url, sheet_name='Recettes')
df_Recette=df_Recette.to_numpy()
df_Demande = pd.read_excel (url, sheet_name='Demande')
df_Demande=df_Demande.to_numpy()
df_transportAmont=pd.read_excel (url, sheet_name='Cout de transport  Amont')
df_transportAmont=df_transportAmont.to_numpy()
df_transportAval=pd.read_excel (url, sheet_name='Cout de transport aval')
df_transportAval=df_transportAval.to_numpy()
df_Capacite_source=pd.read_excel (url, sheet_name='Capacité source')
df_Capacite_source=df_Capacite_source.to_numpy()

#Blenders=[province,1]
SBlenders=[]
Blenders=[]
for i in range(len(df_Blenders)):
    for j in range(df_Blenders[i][1]):
        Blenders.append(df_Blenders[i][0]+"_"+str(j))
for i in range(len(df_Blenders)):
    for j in range(df_Blenders[i][2]):
        SBlenders.append(df_Blenders[i][0]+"_"+str(j))







#Initiation du probleme
prob = Model("OCP problem")



#Process de calcul des constantes
Province_1=getting(0,df_Recette)
Filieres=getting(1,df_Recette)



#Constantes
MP=getting(0,df_MP)
prix_MP=getting(1,df_MP)
Prix_MP=getting(1,df_MP)
Province=getting(0,df_Stockage)
Stock=getting(0,df_Stockage)
Capacite_Stockage=getting(1,df_Stockage)
OCP=["OCP Jorf","OCP SAFI","OCP Port Casablanca","OCP Port Tanger","OCP Port Nador"]
Capacite_Blender=elim_zero(getting(3,df_Blenders))
Capacite_SBlender=elim_zero(getting(4,df_Blenders))
Blenderplus=com_possible(Province,Big_n)
SBlenderplus=com_possible(Province,Big_n_1)
n=len(Blenderplus)
ToutBlenderplus=Blenderplus+SBlenderplus
Capacite_Blenderplus=[12000 for i in range(n)]
Capacite_SBlenderplus=[3000 for i in range(n)]
Capacite_TBlender=Capacite_Blender+Capacite_SBlender
Capacite_TBlenderplus=Capacite_Blenderplus+Capacite_SBlenderplus
Recette=concat_list(Province_1,Filieres)
Recette_avec_ing=making_recette(Recette,df_Recette,MP)
demande={}
for t in range(len(temps)):
    demande[temps[t]]=making_dict(concat_list(getting(0,df_Demande),getting(1,df_Demande)),getting(t+2,df_Demande))
ToutBlender=Blenders+SBlenders
zone=getting(0,df_transportAmont)
Stockplus=com_possible(Province,Big_n)
Capacite_Stockageplus=[3500 for i in range(len(Stockplus))]
coa={}

for i in range(len(zone)):
    m={}
    for j in range(len(OCP)):
        m[OCP[j]]=df_transportAmont[i][j+1]
    coa[zone[i]]=m

cod={}
for i in range(len(zone)):
    m={}
    for j in range(len(zone)):
        m[zone[j]]=df_transportAval[i][j+1]
    cod[zone[i]]=m
def contrainte_min(prob,X,MP,OCP,ToutBlender,W_X,Big_M,blending_min,temps):
    for i in MP:
        for j in OCP:
            for k in ToutBlender:
                for t in temps:
                    prob.addConstr(X[i,j,k,t] <= W_X[i,j,k,t] * Big_M)
                    prob.addConstr(X[i,j,k,t] >= W_X[i,j,k,t] * blending_min)





def cont_activation(prob,Cplus,W_Xplus,MP,OCP,ToutBlenderplus,temps):
    for i in MP:
        for j in OCP:
            for k in ToutBlenderplus:
                for t in temps:
                    prob.addConstr(Cplus[k,t] >= W_Xplus[i,j,k,t])

def cont_activation1(prob,Splus,W_Rplus_2,Recette,Stockplus,ToutBlenderplus):
    for i in Recette:
        for j in Stockplus:
            for k in ToutBlenderplus:
                prob.addConstr(Splus[j] >= W_Rplus_2[i,j,k])
#A revoir
#Déclaration des variables


X=prob.addVars(MP,OCP,ToutBlender,temps,vtype=GRB.CONTINUOUS,lb=0,name="X")
Xplus=prob.addVars(MP,OCP,ToutBlenderplus,temps,vtype=GRB.CONTINUOUS,lb=0,name="Xplus")
R=prob.addVars(MP,Stock,ToutBlender,temps,vtype=GRB.CONTINUOUS,lb=0,name="R")
Rplus_0=prob.addVars(MP,Stock,ToutBlenderplus,temps,vtype=GRB.CONTINUOUS,lb=0,name="Rplus_0")
Rplus_1=prob.addVars(MP,Stockplus,ToutBlender,temps,vtype=GRB.CONTINUOUS,lb=0,name="Rplus_1")
Rplus_2=prob.addVars(MP,Stockplus,ToutBlenderplus,temps,vtype=GRB.CONTINUOUS,lb=0,name="Rplus_2")
Y=prob.addVars(Recette,ToutBlender,Province,temps,vtype=GRB.CONTINUOUS,lb=0,name="Y")
Yplus=prob.addVars(Recette,ToutBlenderplus,Province,temps,vtype=GRB.CONTINUOUS,lb=0,name="Yplus")
A=prob.addVars(MP,OCP,Stock,temps,vtype=GRB.CONTINUOUS,lb=0,name="A")
Aplus=prob.addVars(MP,OCP,Stockplus,temps,vtype=GRB.CONTINUOUS,lb=0,name="Aplus")
B=prob.addVars(Recette,ToutBlender,Stock,temps,vtype=GRB.CONTINUOUS,lb=0,name="B")
Bplus_0=prob.addVars(Recette,ToutBlenderplus,Stock,temps,vtype=GRB.CONTINUOUS,lb=0,name="Bplus_0")
Bplus_1=prob.addVars(Recette,ToutBlender,Stockplus,temps,vtype=GRB.CONTINUOUS,lb=0,name="Bplus_1")
Bplus_2=prob.addVars(Recette,ToutBlenderplus,Stockplus,temps,vtype=GRB.CONTINUOUS,lb=0,name="Bplus_2")
P=prob.addVars(Recette,Stock,Province,temps,vtype=GRB.CONTINUOUS,lb=0,name="P")
Pplus=prob.addVars(Recette,Stockplus,Province,temps,vtype=GRB.CONTINUOUS,lb=0,name="Pplus")


W_X= prob.addVars(MP,OCP,ToutBlender,temps,vtype=GRB.BINARY,lb=0,name="W_X")
W_A= prob.addVars(MP,OCP,Stock,temps,vtype=GRB.BINARY,lb=0,name="W_A")
W_B= prob.addVars(Recette,ToutBlender,Stock,temps,vtype=GRB.BINARY,lb=0,name="W_B")
W_Xplus= prob.addVars(MP,OCP,ToutBlenderplus,temps,vtype=GRB.BINARY,lb=0,name="W_Xplus")
W_Aplus= prob.addVars(MP,OCP,Stockplus,temps,vtype=GRB.BINARY,lb=0,name="W_Aplus")
W_Rplus_0=prob.addVars(MP,Stock,ToutBlenderplus,temps,vtype=GRB.BINARY,lb=0,name="W_Rplus_0")
W_Rplus_1=prob.addVars(MP,Stockplus,ToutBlender,temps,vtype=GRB.BINARY,lb=0,name="W_Rplus_1")
W_Rplus_2=prob.addVars(MP,Stockplus,ToutBlenderplus,temps,vtype=GRB.BINARY,lb=0,name="W_Rplus_2")
W_Yplus=prob.addVars(Recette,ToutBlenderplus,Province,temps,vtype=GRB.BINARY,lb=0,name="W_Yplus")
W_Bplus_0=prob.addVars(Recette,ToutBlenderplus,Stock,temps,vtype=GRB.BINARY,lb=0,name="W_Bplus_0")
W_Bplus_1=prob.addVars(Recette,ToutBlender,Stockplus,temps,vtype=GRB.BINARY,lb=0,name="W_Bplus_1")
W_Bplus_2=prob.addVars(Recette,ToutBlenderplus,Stockplus,temps,vtype=GRB.BINARY,lb=0,name="W_Bplus_2")
W_R=prob.addVars(MP,Stock,ToutBlender,temps,vtype=GRB.BINARY,lb=0,name="W_R")
Cplus=prob.addVars(ToutBlenderplus,temps,vtype=GRB.BINARY,lb=0,name="Cplus")
Splus=prob.addVars(Stockplus,temps,vtype=GRB.BINARY,lb=0,name="Splus")
res_MP=prob.addVars(MP,Stock,temps,vtype=GRB.CONTINUOUS,lb=0,name="res_MP")
res_PF=prob.addVars(Recette,Stock,temps,vtype=GRB.CONTINUOUS,lb=0,name="res_PF")
res_MP_plus=prob.addVars(MP,Stockplus,temps,vtype=GRB.CONTINUOUS,lb=0,name="res_MP_plus")
res_PF_plus=prob.addVars(Recette,Stockplus,temps,vtype=GRB.CONTINUOUS,lb=0,name="res_PF_plus")
CplusF=prob.addVars(ToutBlenderplus,vtype=GRB.BINARY,lb=0,name="CplusF")
SplusF=prob.addVars(Stockplus,vtype=GRB.BINARY,lb=0,name="SplusF")




#Contraintes

for i in ToutBlenderplus:
    for t in temps:
        prob.addConstr(CplusF[i]>=Cplus[i,t])

for i in Stockplus:
    for t in temps:
        prob.addConstr(SplusF[i]>=Splus[i,t])




cont_activation(prob,Cplus,W_Xplus,MP,OCP,ToutBlenderplus,temps)
cont_activation(prob,Cplus,W_Rplus_0,MP,Stock,ToutBlenderplus,temps)
cont_activation(prob,Cplus,W_Rplus_2,MP,Stockplus,ToutBlenderplus,temps)
cont_activation(prob,Splus,W_Aplus,MP,OCP,Stockplus,temps)
cont_activation(prob,Splus,W_Bplus_1,Recette,ToutBlender,Stockplus,temps)
cont_activation(prob,Splus,W_Bplus_2,Recette,ToutBlenderplus,Stockplus,temps)



#Contrainte mini
contrainte_min(prob,R,MP,Stock,ToutBlender,W_R,Big_M,blending_min,temps)
contrainte_min(prob,X,MP,OCP,ToutBlender,W_X,Big_M,blending_min,temps)
contrainte_min(prob,Xplus,MP,OCP,ToutBlenderplus,W_Xplus,Big_M,blending_min,temps)
contrainte_min(prob,A,MP,OCP,Stock,W_A,Big_M,stockage_min,temps)
contrainte_min(prob,Aplus,MP,OCP,Stockplus,W_Aplus,Big_M,stockage_min,temps)
contrainte_min(prob,B,Recette,ToutBlender,Stock,W_B,Big_M,stockage_min,temps)
contrainte_min(prob,Bplus_1,Recette,ToutBlender,Stockplus,W_Bplus_1,Big_M,stockage_min,temps)
contrainte_min(prob,Bplus_0,Recette,ToutBlenderplus,Stock,W_Bplus_0,Big_M,stockage_min,temps)
contrainte_min(prob,Bplus_2,Recette,ToutBlenderplus,Stockplus,W_Bplus_2,Big_M,stockage_min,temps)
contrainte_min(prob,Yplus,Recette,ToutBlenderplus,Province,W_Yplus,Big_M,0,temps)
contrainte_min(prob,Rplus_0,MP,Stock,ToutBlenderplus,W_Rplus_0,Big_M,0,temps)
contrainte_min(prob,Rplus_1,MP,Stockplus,ToutBlender,W_Rplus_1,Big_M,0,temps)
contrainte_min(prob,Rplus_2,MP,Stockplus,ToutBlenderplus,W_Rplus_2,Big_M,0,temps)





for k in range(len(ToutBlender)):
    prob.addConstr(quicksum(X[i,j,ToutBlender[k],t] for t in temps for i in MP for j in OCP)
                   +quicksum(R[i,j,ToutBlender[k],t] for t in temps for i in MP for j in Stock)
                   +quicksum(Rplus_1[i,j,ToutBlender[k],t] for t in temps for i in MP for j in Stockplus)<= Capacite_TBlender[k])




for k in range(len(ToutBlenderplus)):

    prob.addConstr(quicksum(Xplus[i,j,ToutBlenderplus[k],t] for t in temps for i in MP for j in OCP)
                   + quicksum(Rplus_0[i,j,ToutBlenderplus[k],t] for t in temps for i in MP for j in Stock)
                   + quicksum(Rplus_2[i,j,ToutBlenderplus[k],t] for t in temps for i in MP for j in Stockplus) <=
                   Capacite_TBlenderplus[k])

for k in range(len(Stock)):

    prob.addConstr(quicksum(A[i,j,Stock[k],t] for t in temps for i in MP for j in OCP)
                   +quicksum(B[i,j,Stock[k],t] for t in temps for i in Recette for j in ToutBlender)
                   +quicksum(Bplus_0[i,j,Stock[k],t] for t in temps for i in Recette for j in ToutBlenderplus)
                   +quicksum(res_MP[i,Stock[k],t] for t in temps for i in MP)
                   +quicksum(res_PF[i,Stock[k],t] for t in temps for i in Recette)<= Capacite_Stockage[k]
                   )


for k in range(len(Stockplus)):


    prob.addConstr(quicksum(Aplus[i,j,Stockplus[k],t] for t in temps for i in MP for j in OCP)
                   + quicksum(Bplus_1[i,j,Stockplus[k],t] for t in temps for i in Recette for j in ToutBlender)
                   + quicksum(Bplus_2[i,j,Stockplus[k],t] for t in temps for i in Recette for j in ToutBlenderplus)
                   + quicksum(res_MP_plus[i,Stockplus[k],t] for t in temps for i in MP)
                   + quicksum(res_PF_plus[i,Stockplus[k],t] for t in temps for i in Recette) <= Capacite_Stockageplus[k]
                   )

for k in range(len(ToutBlender)):
    prob.addConstr(quicksum(X[i,j,ToutBlender[k],t] for t in temps for i in MP for j in OCP)
                   +quicksum(R[i,j,ToutBlender[k],t] for t in temps for i in MP for j in Stock)
                   +quicksum(Rplus_1[i,j,ToutBlender[k],t] for t in temps for i in MP for j in Stockplus)
                   ==quicksum(Y[i,ToutBlender[k],j,t] for t in temps for i in Recette for j in Province)
                   +quicksum(B[i,ToutBlender[k],j,t] for t in temps for i in Recette for j in Stock)
                   +quicksum(Bplus_1[i,ToutBlender[k],j,t] for t in temps for i in Recette for j in Stockplus)
                   )

for k in range(len(ToutBlenderplus)):

    prob.addConstr(quicksum(Xplus[i,j,ToutBlenderplus[k],t] for t in temps for i in MP for j in OCP)
                   + quicksum(Rplus_0[i,j,ToutBlenderplus[k],t] for t in temps for i in MP for j in Stock)
                   + quicksum(Rplus_2[i,j,ToutBlenderplus[k],t] for t in temps for i in MP for j in Stockplus)
                   == quicksum(Yplus[i,ToutBlenderplus[k],j,t] for t in temps for i in Recette for j in Province)
                   + quicksum(Bplus_0[i,ToutBlenderplus[k],j,t] for t in temps for i in Recette for j in Stock)
                   + quicksum(Bplus_2[i,ToutBlenderplus[k],j,t] for t in temps for i in Recette for j in Stockplus)
                   )





#demande satistfaite:
for i in Recette:
    for k in Province:
        for t in temps:
            if i.split("_")[0] == k:
                prob.addConstr(quicksum(Y[i,j,k,t] for j in ToutBlender)+quicksum(Yplus[i,j,k,t] for j in ToutBlenderplus)
                               +quicksum(P[i,j,k,t] for j in Stock)+quicksum(Pplus[i,j,k,t] for j in Stockplus)== demande[t][i])


#Les blenders doivent recevoir de quoi faire les recettes
for i in MP:
    for k in ToutBlender:
        for t in temps:
            prob.addConstr(quicksum(X[i,j,k,t] for j in OCP)
                           +quicksum(R[i,j,k,t] for j in Stock)
                           +quicksum(Rplus_1[i,j,k,t] for j in Stockplus)
                           ==quicksum(Recette_avec_ing[a][i] * Y[a,k,j,t] for j in Province for a in Recette)
                           +quicksum(Recette_avec_ing[a][i] * B[a,k,j,t] for j in Stock for a in Recette)
                           +quicksum(Recette_avec_ing[a][i] * Bplus_1[a,k,j,t] for j in Stockplus for a in Recette))

for i in MP:
    for k in ToutBlenderplus:
        for t in temps:
            prob.addConstr(quicksum(Xplus[i, j, k, t] for j in OCP)
                           + quicksum(Rplus_0[i, j, k, t] for j in Stock)
                           + quicksum(Rplus_2[i,j,k,t] for j in Stockplus)
                           == quicksum(Recette_avec_ing[a][i] * Yplus[a, k, j, t] for j in Province for a in Recette)
                           + quicksum(Recette_avec_ing[a][i] * Bplus_0[a, k, j, t] for j in Stock for a in Recette)
                           + quicksum(Recette_avec_ing[a][i] * Bplus_2[a, k, j, t] for j in Stockplus for a in Recette))

#Les stocks ne creent rien
for i in MP:
    for j in Stock:
        for t in range(1,len(temps)):
            prob.addConstr(quicksum(R[i,j,k,temps[t]] for k in ToutBlender)
                           +quicksum(Rplus_0[i,j,k,temps[t]] for k in ToutBlenderplus)
                           <=quicksum(A[i,k,j,temps[t]] for k in OCP)+res_MP[i,j,temps[t-1]])
        prob.addConstr(quicksum(R[i, j, k, 2] for k in ToutBlender)
                       + quicksum(Rplus_0[i, j, k, 2] for k in ToutBlenderplus)
                       <= quicksum(A[i, k, j, 2] for k in OCP) )

for i in MP:
    for j in Stockplus:
        for t in range(1,len(temps)):
            prob.addConstr(quicksum(Rplus_1[i,j,k,temps[t]] for k in ToutBlender)
                           + quicksum(Rplus_2[i,j,k,temps[t]] for k in ToutBlenderplus)
                           <= quicksum(Aplus[i,k,j,temps[t]] for k in OCP) + res_MP_plus[i,j,temps[t-1]])
        prob.addConstr(quicksum(Rplus_1[i, j, k, 2] for k in ToutBlender)
                       + quicksum(Rplus_2[i, j, k, 2] for k in ToutBlenderplus)
                       <= quicksum(Aplus[i, k, j, 2] for k in OCP) )

for i in Recette:
    for j in Stock:
        for t in range(1,len(temps)):
            prob.addConstr(quicksum(P[i,j,k,temps[t]] for k in Province)<=
                           quicksum(B[i,k,j,temps[t]] for k in ToutBlender)
                           +quicksum(Bplus_0[i,k,j,temps[t]] for k in ToutBlenderplus)
                           +res_PF[i,j,temps[t-1]])
        prob.addConstr(quicksum(P[i, j, k, 2] for k in Province) <=
                       quicksum(B[i, k, j, 2] for k in ToutBlender)
                       + quicksum(Bplus_0[i, k, j, 2] for k in ToutBlenderplus)
                       )

for i in Recette:
    for j in Stockplus:
        for t in range(1,len(temps)):
            prob.addConstr(quicksum(Pplus[i,j,k,temps[t]] for k in Province) <=
                           quicksum(Bplus_1[i,k,j,temps[t]] for k in ToutBlender)
                           + quicksum(Bplus_2[i,k,j,temps[t]] for k in ToutBlenderplus)
                           + res_PF_plus[i,j,temps[t-1]])
        prob.addConstr(quicksum(Pplus[i, j, k, 2] for k in Province) <=
                       quicksum(Bplus_1[i, k, j, 2] for k in ToutBlender)
                       + quicksum(Bplus_2[i, k, j, 2] for k in ToutBlenderplus)
                       )


#Contrainte de supply:
for i in range(len(MP)):
    for j in range(len(OCP)):
        for t in temps:
            prob.addConstr(quicksum(X[MP[i],OCP[j],k,t] for k in ToutBlender)
                           +quicksum(Xplus[MP[i],OCP[j],k,t] for k in ToutBlenderplus)
                           +quicksum(A[MP[i],OCP[j],k,t] for k in Stock)
                           +quicksum(Aplus[MP[i],OCP[j],k,t] for k in Stockplus)
                           <=df_Capacite_source[i][j + 1]
                           )

#Step 3 residus:

for i in MP:
    for j in Stock:
        prob.addConstr(res_MP[i,j,2]==quicksum(A[i,k,j,2] for k in OCP)
                       -quicksum(R[i,j,k,2] for k in ToutBlender)
                       -quicksum(Rplus_0[i,j,k,2] for k in ToutBlenderplus))
        for t in range(1,len(temps)):
            prob.addConstr(res_MP[i,j,temps[t]] == res_MP[i,j,temps[t-1]]
                           +quicksum(A[i,k,j,temps[t]] for k in OCP)
                           -quicksum(R[i,j,k,temps[t]] for k in ToutBlender)
                           -quicksum(Rplus_0[i,j,k,temps[t]] for k in ToutBlenderplus))
    for j in Stockplus:

        prob.addConstr(res_MP_plus[i,j,2] == quicksum(Aplus[i,k,j,2] for k in OCP)
                       - quicksum(Rplus_1[i,j,k,2] for k in ToutBlender)
                       - quicksum(Rplus_2[i,j,k,2] for k in ToutBlenderplus))
        for t in range(1, len(temps)):
            prob.addConstr(res_MP_plus[i,j,temps[t]] == res_MP_plus[i,j,temps[t - 1]]
                           + quicksum(Aplus[i,k,j,temps[t]] for k in OCP)
                           - quicksum(Rplus_1[i,j,k,temps[t]] for k in ToutBlender)
                           - quicksum(Rplus_2[i,j,k,temps[t]] for k in ToutBlenderplus))

for i in Recette:
    for j in Stock:
        prob.addConstr(res_PF[i,j,2] == quicksum(B[i,k,j,2] for k in ToutBlender)
                       +quicksum(Bplus_0[i,k,j,2] for k in ToutBlenderplus)
                       - quicksum(P[i,j,k,2] for k in Province)
                       )
        for t in range(1,len(temps)):
            prob.addConstr(res_PF[i,j,temps[t]] == res_PF[i,j,temps[t-1]]
                           +quicksum(B[i,k,j,temps[t]] for k in ToutBlender)
                           +quicksum(Bplus_0[i,k,j,temps[t]] for k in ToutBlenderplus)
                           -quicksum(P[i,j,k,temps[t]] for k in Province))
    for j in Stockplus:
        prob.addConstr(res_PF_plus[i,j,2] == quicksum(Bplus_1[i,k,j,2] for k in ToutBlender)
                       + quicksum(Bplus_2[i,k,j,2] for k in ToutBlenderplus)
                       - quicksum(Pplus[i,j,k,2] for k in Province)
                       )
        for t in range(1, len(temps)):
            prob.addConstr(res_PF_plus[i,j,temps[t]] == res_PF_plus[i,j,temps[t - 1]]
                           + quicksum(Bplus_1[i,k,j,temps[t]] for k in ToutBlender)
                           + quicksum(Bplus_2[i,k,j,temps[t]] for k in ToutBlenderplus)
                           - quicksum(Pplus[i,j,k,temps[t]] for k in Province)
                           )



#Objective function


prob.setObjective(
quicksum(Cout_stockage_building*SplusF[k] for k in Stockplus)
+quicksum(Cout_sblender*CplusF[k] for k in SBlenderplus)
+quicksum(Cout_blender*CplusF[k] for k in Blenderplus)
+quicksum(Cout_de_stockage_Engrais*Bplus_2[i,j,k,t] for t in temps for i in Recette for j in ToutBlenderplus for k in Stockplus)
+quicksum(Cout_de_stockage_Engrais*Bplus_1[i,j,k,t] for t in temps for i in Recette for j in ToutBlender for k in Stockplus)
+quicksum(Cout_de_stockage_Engrais*Bplus_0[i,j,k,t] for t in temps for i in Recette for j in ToutBlenderplus for k in Stock)
+quicksum(Cout_de_stockage_Engrais*B[i,j,k,t] for t in temps for i in Recette for j in ToutBlender for k in Stock)
+quicksum(Cout_de_stockage_MP*Aplus[i,j,k,t] for t in temps for i in MP for j in OCP for k in Stockplus)
+quicksum(Cout_de_stockage_MP*A[i,j,k,t] for t in temps for i in MP for j in OCP for k in Stock)
+quicksum(Cout_smart_blending*Rplus_2[i,j,k,t] for t in temps for i in MP for j in Stockplus for k in SBlenderplus)
+quicksum(Cout_smart_blending*Rplus_1[i,j,k,t] for t in temps for i in MP for j in Stockplus for k in SBlenders)
+quicksum(Cout_smart_blending*Rplus_0[i,j,k,t] for t in temps for i in MP for j in Stock for k in SBlenderplus)
+quicksum(Cout_smart_blending*R[i,j,k,t] for t in temps for i in MP for j in Stock for k in SBlenders)
+quicksum(Cout_blending*Rplus_2[i,j,k,t] for t in temps for i in MP for j in Stockplus for k in Blenderplus)
+quicksum(Cout_blending*Rplus_1[i,j,k,t] for t in temps for i in MP for j in Stockplus for k in Blenders)
+quicksum(Cout_blending*Rplus_0[i,j,k,t] for t in temps for i in MP for j in Stock for k in Blenderplus)
+quicksum(Cout_blending*R[i,j,k,t] for t in temps for i in MP for j in Stock for k in Blenders)
+quicksum(Cout_smart_blending*Xplus[i,j,k,t] for t in temps for i in MP for j in OCP for k in SBlenderplus)
+quicksum(Cout_smart_blending*X[i,j,k,t] for t in temps for i in MP for j in OCP for k in SBlenders)
+quicksum(Cout_blending*Xplus[i,j,k,t] for t in temps for i in MP for j in OCP for k in Blenderplus)
+quicksum(Cout_blending*X[i,j,k,t] for t in temps for i in MP for j in OCP for k in Blenders)
+quicksum(coa[k.split("_")[0]][j]*Aplus[i,j,k,t] for t in temps for i in MP for j in OCP for k in Stockplus)
+quicksum(coa[k.split("_")[0]][j]*A[i,j,k,t] for t in temps for i in MP for j in OCP for k in Stock)
+quicksum(cod[k.split("_")[0]][j.split("_")[0]]*Bplus_2[i,j,k,t] for t in temps for i in Recette for j in ToutBlenderplus for k in Stockplus)
+quicksum(cod[k.split("_")[0]][j.split("_")[0]]*Bplus_1[i,j,k,t] for t in temps for i in Recette for j in ToutBlender for k in Stockplus)
+quicksum(cod[k][j.split("_")[0]]*Bplus_0[i,j,k,t] for t in temps for i in Recette for j in ToutBlenderplus for k in Stock)
+quicksum(cod[k][j.split("_")[0]]*B[i,j,k,t] for t in temps for i in Recette for j in ToutBlender for k in Stock)
+quicksum(cod[k][j.split("_")[0]]*Yplus[i,j,k,t] for t in temps for i in Recette for j in ToutBlenderplus for k in Province)
+quicksum(cod[k][j.split("_")[0]]*Y[i,j,k,t] for t in temps for i in Recette for j in ToutBlender for k in Province)
+quicksum(coa[k.split("_")[0]][j]*Xplus[i,j,k,t] for t in temps for i in MP for j in OCP for k in ToutBlenderplus)
+quicksum(coa[k.split("_")[0]][j]*X[i,j,k,t] for t in temps for i in MP for j in OCP for k in ToutBlender)
+quicksum(cod[j.split("_")[0]][k.split("_")[0]]*R[i,j,k,t] for t in temps for i in MP for j in Stock for k in ToutBlender)
+quicksum(cod[j.split("_")[0]][k.split("_")[0]]*Rplus_0[i,j,k,t] for t in temps for i in MP for j in Stock for k in ToutBlenderplus)
+quicksum(cod[j.split("_")[0]][k.split("_")[0]]*Rplus_1[i,j,k,t] for t in temps for i in MP for j in Stockplus for k in ToutBlender)
+quicksum(cod[j.split("_")[0]][k.split("_")[0]]*Rplus_2[i,j,k,t] for t in temps for i in MP for j in Stockplus for k in ToutBlenderplus)
+quicksum(cod[j.split("_")[0]][k.split("_")[0]]*P[i,j,k,t] for t in temps for i in Recette for j in Stock for k in Province)
+quicksum(cod[j.split("_")[0]][k.split("_")[0]]*Pplus[i,j,k,t] for t in temps for i in Recette for j in Stockplus for k in Province)


,GRB.MINIMIZE)


prob.optimize()


row=2
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute
sheet1 = wb.active




# iterating through content list
for i in range(len(OCP)):
    for j in range(len(MP)):
        for k in range(len(Stock)):
            for t in range(len(temps)):
                if A[MP[j],OCP[i],Stock[k],temps[t]].X!=0  :
                    c1 = sheet1.cell(row=row , column=1)
                    c1.value = OCP[i]
                    c2 = sheet1.cell(row=row , column=2)
                    c2.value = MP[j]
                    c3 = sheet1.cell(row=row , column=3)
                    c3.value = Stock[k].split("_")[0]
                    c4 = sheet1.cell(row=row , column=4)
                    c4.value = "Q" + str(temps[t])
                    c5 = sheet1.cell(row=row , column=5)
                    c5.value = A[MP[j], OCP[i], Stock[k], temps[t]].X
                    c6 = sheet1.cell(row=row, column=6)
                    c6.value = prix_MP[j]*A[MP[j], OCP[i], Stock[k], temps[t]].X


                    row += 1


wb.save("C:\\Users\\hp\\Desktop\\optimisation\\demo.xlsx")

row=2
wb1 = openpyxl.Workbook()
sheet2 = wb1.active
for i in range(len(Province)):
    for j in range(len(temps)):
        c1 = sheet2.cell(row=row, column=1)
        c1.value = OCP[i]
        c2 = sheet2.cell(row=row, column=2)
        c2.value = MP[j]
        c3 = sheet2.cell(row=row, column=3)
        c3.value = Stock[k].split("_")[0]
        row+=1




wb1.save("C:\\Users\\hp\\Desktop\\optimisation\\demo_1.xlsx")
